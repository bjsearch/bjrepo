"""
Excel 파일에서 보험 가입 정보를 읽고 분석 리포트 데이터로 변환.

지원 형식:
1. 구조화된 형식:
   - 시트 "고객정보": A1=고객명, B1=성별, C1=생년월일, D1=분석기준일, E1=담당자
   - 시트 "보험상품": 헤더행 + 데이터 (보험사, 상품명, 계약일, 월보험료, 총보험료 등)

2. 기타 형식:
   - 첫 번째 시트에서 자동 감지하여 파싱

pandas를 사용하여 손상된 포맷의 Excel도 읽을 수 있습니다.
"""
from __future__ import annotations

import re
import math
from datetime import date, datetime
from dataclasses import dataclass

from .parser import Customer, DetailItem, ParsedReport


@dataclass
class ExcelParseResult:
    """Excel 파일 파싱 결과."""
    customer_name: str
    gender: str | None
    birth_date: date | None
    basis_date: date | None
    handler: str
    insurance_products: list[dict]

    def to_parsed_report(self) -> ParsedReport:
        """PDF 파서와 호환되는 ParsedReport 형식으로 변환."""
        age = None
        if self.birth_date:
            today = datetime.now().date()
            age = today.year - self.birth_date.year
            if (today.month, today.day) < (self.birth_date.month, self.birth_date.day):
                age -= 1

        customer = Customer(
            name=self.customer_name,
            rrn_masked="***-****-**",
            birth_date=self.birth_date,
            gender=self.gender,
            age_insurance=age,
            handler=self.handler,
            basis_date=self.basis_date or datetime.now().date()
        )

        detail_items = []
        for product in self.insurance_products:
            coverages = product.get("coverages", [])
            category = ""
            if coverages:
                category = coverages[0].get("name", "")
            else:
                category = product.get("product_name", "")[:20]

            total_premium = product.get("total_premium", 0)
            if total_premium > 0:
                amount_man = total_premium // 10000
            else:
                monthly_premium = product.get("monthly_premium", 0)
                amount_man = (monthly_premium * 12) // 10000

            contract_date_str = product.get("contract_date", "")
            if contract_date_str:
                if isinstance(contract_date_str, date):
                    start_str = contract_date_str.strftime("%Y-%m-%d")
                else:
                    start_str = str(contract_date_str)
            else:
                start_str = ""

            end_str = product.get("contract_end", "") or "9999-12-31"
            if isinstance(end_str, date):
                end_str = end_str.strftime("%Y-%m-%d")

            detail_item = DetailItem(
                company=product.get("company", ""),
                product=product.get("product_name", ""),
                start=start_str,
                end=str(end_str),
                pay_years=None,
                pay_method="월납",
                premium_won=product.get("monthly_premium", 0),
                rider_name="",
                category=category,
                amount_man=int(amount_man) if amount_man > 0 else 0,
                status=""
            )
            detail_items.append(detail_item)

        return ParsedReport(customer=customer, detail_items=detail_items)


def parse_excel(file_path: str) -> ExcelParseResult:
    """Excel 파일을 파싱해서 보험 정보를 추출한다."""
    try:
        import pandas as pd
    except ImportError:
        raise ReportParseError("pandas 패키지가 필요합니다")

    try:
        # data_only=True로 값만 읽음 (포맷/스타일 무시)
        excel_file = pd.ExcelFile(file_path, engine='openpyxl')
    except Exception as e:
        # 스타일 파싱 오류 무시하고 재시도
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            # pandas가 실패했으면 수동 파싱으로 진행
            return _parse_excel_manual(file_path, sheet_names)
        except:
            raise ReportParseError(f"Excel 파일을 읽을 수 없습니다: {str(e)[:100]}")

    customer_name = "미입력"
    gender = None
    birth_date = None
    basis_date = None
    handler = "사용자"

    # "고객정보" 시트가 있으면 읽기
    if "고객정보" in excel_file.sheet_names:
        try:
            df = pd.read_excel(file_path, sheet_name="고객정보", header=None, engine='openpyxl')
        except:
            df = None

        if df is not None and not df.empty:
            customer_name = _safe_value(df.iloc[0, 0]) or "미입력"
            gender = _safe_value(df.iloc[0, 1]) if len(df.columns) > 1 else None
            birth_str = _safe_value(df.iloc[0, 2]) if len(df.columns) > 2 else None
            basis_str = _safe_value(df.iloc[0, 3]) if len(df.columns) > 3 else None
            handler_val = _safe_value(df.iloc[0, 4]) if len(df.columns) > 4 else None

            if birth_str:
                birth_date = _parse_date(birth_str)
            if basis_str:
                basis_date = _parse_date(basis_str)
            if handler_val:
                handler = handler_val

    # "보험상품" 시트가 있으면 읽기
    insurance_products = []
    if "보험상품" in excel_file.sheet_names:
        try:
            df = pd.read_excel(file_path, sheet_name="보험상품", header=None, engine='openpyxl')
        except:
            df = None

        if df is not None and not df.empty:
            insurance_products = _extract_products_from_df(df)

    # 보험상품이 없으면, 첫 번째 시트에서 자동 감지 시도
    if not insurance_products and len(excel_file.sheet_names) > 0:
        sheet_name = excel_file.sheet_names[0]
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
        except:
            df = None

        if df is not None and not df.empty:
            insurance_products = _extract_products_from_df(df)

    # 보험상품이 여전히 없으면 에러
    if not insurance_products:
        return ExcelParseResult(
            customer_name=customer_name,
            gender=gender,
            birth_date=birth_date,
            basis_date=basis_date,
            handler=handler,
            insurance_products=[{
                "company": "미입력",
                "product_name": "데이터 없음",
                "contract_date": "",
                "monthly_premium": 0,
                "total_premium": 0,
                "coverages": [{"name": "지원되지 않는 Excel 형식", "amount": 0}],
                "contract_end": "9999-12-31"
            }]
        )

    return ExcelParseResult(
        customer_name=customer_name,
        gender=gender,
        birth_date=birth_date,
        basis_date=basis_date,
        handler=handler,
        insurance_products=insurance_products
    )


def _extract_products_from_df(df) -> list[dict]:
    """DataFrame에서 보험상품 정보를 추출한다."""
    products = []

    # 첫 행이 헤더로 보이는 경우 (빈 첫 행 제외)
    start_idx = 0
    for idx in range(len(df)):
        first_col = _safe_value(df.iloc[idx, 0])
        if first_col:
            start_idx = idx
            break

    # 시작 행부터 읽기
    for idx in range(start_idx + 1, len(df)):
        try:
            company = _safe_value(df.iloc[idx, 0])
            if not company:
                break

            product_name = _safe_value(df.iloc[idx, 1]) if len(df.columns) > 1 else None
            if not product_name:
                continue

            product = {
                "company": company,
                "product_name": product_name,
                "contract_date": _safe_value(df.iloc[idx, 2]) if len(df.columns) > 2 else "",
                "monthly_premium": _parse_number(_safe_value(df.iloc[idx, 3])) if len(df.columns) > 3 else 0,
                "total_premium": _parse_number(_safe_value(df.iloc[idx, 4])) if len(df.columns) > 4 else 0,
                "remaining_premium": _parse_number(_safe_value(df.iloc[idx, 5])) if len(df.columns) > 5 else 0,
                "coverages": _parse_coverages(_safe_value(df.iloc[idx, 6])) if len(df.columns) > 6 else [],
                "contract_end": _safe_value(df.iloc[idx, 7]) if len(df.columns) > 7 else ""
            }
            products.append(product)
        except Exception:
            continue

    return products


def _safe_value(val) -> str | None:
    """pandas DataFrame 셀 값을 안전하게 추출한다."""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    val_str = str(val).strip()
    return val_str if val_str else None


def _parse_excel_manual(file_path: str, sheet_names: list) -> ExcelParseResult:
    """ZIP 기반 Excel 파싱 (스타일 무시)"""
    import zipfile
    import xml.etree.ElementTree as ET
    from io import BytesIO

    try:
        # Excel을 ZIP으로 열기 (xlsx는 ZIP 형식)
        with zipfile.ZipFile(file_path, 'r') as zf:
            # 워크북 정보 읽기
            workbook_xml = zf.read('xl/workbook.xml')
            root = ET.fromstring(workbook_xml)

            # 네임스페이스
            ns = {'': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

            # 시트 목록
            sheets = {}
            for sheet_elem in root.findall('.//sheet', ns):
                name = sheet_elem.get('name', '')
                rid = sheet_elem.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id', '')
                sheets[name] = rid

            # 관계 정보 (rel ID -> 파일명 매핑)
            rels_xml = zf.read('xl/_rels/workbook.xml.rels')
            rels_root = ET.fromstring(rels_xml)
            sheet_files = {}
            for rel in rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                rel_id = rel.get('Id', '')
                target = rel.get('Target', '')
                if target.endswith('.xml'):
                    sheet_files[rel_id] = target

            # 고객 정보와 보험상품 추출
            customer_name = "미입력"
            gender = None
            birth_date = None
            basis_date = None
            handler = "사용자"
            insurance_products = []

            # 각 시트 읽기
            for sheet_name in sheets:
                rid = sheets[sheet_name]
                if rid in sheet_files:
                    sheet_path = 'xl/' + sheet_files[rid]
                    try:
                        sheet_xml = zf.read(sheet_path)
                        sheet_root = ET.fromstring(sheet_xml)

                        # 셀 데이터 추출
                        cells = {}
                        for c in sheet_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                            r = c.get('r', '')  # 셀 참조 (A1, B2 등)
                            v_elem = c.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                            value = v_elem.text if v_elem is not None else None
                            if value:
                                cells[r] = value

                        # 고객정보 시트인 경우
                        if sheet_name == "고객정보" or sheet_name == "고객 정보":
                            customer_name = cells.get('A1', "미입력")
                            gender = cells.get('B1')
                            birth_str = cells.get('C1')
                            basis_str = cells.get('D1')
                            handler_val = cells.get('E1')

                            if birth_str:
                                birth_date = _parse_date(birth_str)
                            if basis_str:
                                basis_date = _parse_date(basis_str)
                            if handler_val:
                                handler = handler_val

                        # 보험상품 시트인 경우
                        elif sheet_name == "보험상품":
                            row = 2
                            while True:
                                company = cells.get(f'A{row}')
                                if not company:
                                    break

                                product_name = cells.get(f'B{row}')
                                if not product_name:
                                    row += 1
                                    continue

                                product = {
                                    "company": company,
                                    "product_name": product_name,
                                    "contract_date": cells.get(f'C{row}', ''),
                                    "monthly_premium": _parse_number(cells.get(f'D{row}')),
                                    "total_premium": _parse_number(cells.get(f'E{row}')),
                                    "remaining_premium": _parse_number(cells.get(f'F{row}')),
                                    "coverages": _parse_coverages(cells.get(f'G{row}')),
                                    "contract_end": cells.get(f'H{row}', '')
                                }
                                insurance_products.append(product)
                                row += 1
                    except:
                        continue

        return ExcelParseResult(
            customer_name=customer_name,
            gender=gender,
            birth_date=birth_date,
            basis_date=basis_date,
            handler=handler,
            insurance_products=insurance_products
        )

    except Exception as e:
        raise ReportParseError(f"Excel 파일을 읽을 수 없습니다: {str(e)[:100]}")

    customer_name = "미입력"
    gender = None
    birth_date = None
    basis_date = None
    handler = "사용자"

    # 고객정보 시트 읽기
    if "고객정보" in wb.sheetnames:
        ws = wb["고객정보"]
        customer_name = _get_cell_value_openpyxl(ws, "A1") or "미입력"
        gender = _get_cell_value_openpyxl(ws, "B1")
        birth_str = _get_cell_value_openpyxl(ws, "C1")
        basis_str = _get_cell_value_openpyxl(ws, "D1")
        handler_val = _get_cell_value_openpyxl(ws, "E1")

        if birth_str:
            birth_date = _parse_date(birth_str)
        if basis_str:
            basis_date = _parse_date(basis_str)
        if handler_val:
            handler = handler_val

    # 보험상품 시트 읽기
    insurance_products = []
    if "보험상품" in wb.sheetnames:
        ws = wb["보험상품"]
        for row_idx in range(2, min(ws.max_row + 1, 1000)):
            company = _get_cell_value_openpyxl(ws, f"A{row_idx}")
            if not company:
                break

            product_name = _get_cell_value_openpyxl(ws, f"B{row_idx}") or ""
            if not product_name:
                continue

            product = {
                "company": company,
                "product_name": product_name,
                "contract_date": _get_cell_value_openpyxl(ws, f"C{row_idx}") or "",
                "monthly_premium": _parse_number(_get_cell_value_openpyxl(ws, f"D{row_idx}")),
                "total_premium": _parse_number(_get_cell_value_openpyxl(ws, f"E{row_idx}")),
                "remaining_premium": _parse_number(_get_cell_value_openpyxl(ws, f"F{row_idx}")) if ws.max_column >= 6 else 0,
                "coverages": _parse_coverages(_get_cell_value_openpyxl(ws, f"G{row_idx}")) if ws.max_column >= 7 else [],
                "contract_end": _get_cell_value_openpyxl(ws, f"H{row_idx}") if ws.max_column >= 8 else ""
            }
            insurance_products.append(product)

    wb.close()

    if not insurance_products:
        # 첫 시트에서 자동 감지 시도
        if len(wb.sheetnames) > 0:
            sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            for row_idx in range(2, min(ws.max_row + 1, 1000)):
                company = _get_cell_value_openpyxl(ws, f"A{row_idx}")
                if not company:
                    break
                product_name = _get_cell_value_openpyxl(ws, f"B{row_idx}") or ""
                if not product_name:
                    continue
                product = {
                    "company": company,
                    "product_name": product_name,
                    "contract_date": _get_cell_value_openpyxl(ws, f"C{row_idx}") or "",
                    "monthly_premium": _parse_number(_get_cell_value_openpyxl(ws, f"D{row_idx}")),
                    "total_premium": _parse_number(_get_cell_value_openpyxl(ws, f"E{row_idx}")),
                    "remaining_premium": _parse_number(_get_cell_value_openpyxl(ws, f"F{row_idx}")) if ws.max_column >= 6 else 0,
                    "coverages": _parse_coverages(_get_cell_value_openpyxl(ws, f"G{row_idx}")) if ws.max_column >= 7 else [],
                    "contract_end": _get_cell_value_openpyxl(ws, f"H{row_idx}") if ws.max_column >= 8 else ""
                }
                insurance_products.append(product)

    return ExcelParseResult(
        customer_name=customer_name,
        gender=gender,
        birth_date=birth_date,
        basis_date=basis_date,
        handler=handler,
        insurance_products=insurance_products
    )


def _get_cell_value_openpyxl(ws, cell_ref: str) -> str | None:
    """openpyxl 셀 값을 안전하게 추출한다."""
    try:
        cell = ws[cell_ref]
        value = cell.value
        if value is None:
            return None
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        return str(value).strip() if str(value).strip() else None
    except:
        return None


def _get_cell_value(ws, cell_ref: str) -> str | None:
    """셀 값을 안전하게 가져온다."""
    try:
        cell = ws[cell_ref]
        value = cell.value
        if value is None:
            return None
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        return str(value).strip()
    except:
        return None


def _parse_date(date_str: str | None) -> date | None:
    """다양한 형식의 날짜 문자열을 파싱한다."""
    if not date_str:
        return None

    date_str = date_str.strip()

    # 이미 date 객체인 경우
    if isinstance(date_str, date):
        return date_str

    # 다양한 형식 시도
    formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%Y년 %m월 %d일",
        "%m/%d/%Y",
        "%m-%d-%Y",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue

    return None


def _parse_number(value: str | int | float | None) -> int:
    """숫자 문자열을 정수로 파싱한다."""
    if value is None:
        return 0

    if isinstance(value, (int, float)):
        return int(value)

    value_str = str(value).strip()
    # 쉼표, 공백 제거
    value_str = re.sub(r'[,\s]', '', value_str)
    # 한글 단위 제거
    value_str = re.sub(r'[만원]', '', value_str)

    try:
        return int(float(value_str))
    except ValueError:
        return 0


def _parse_coverages(coverage_str: str | None) -> list[dict]:
    """보장내용 문자열을 파싱한다. 예: "상해보장 1000만원, 질병보장 500만원" """
    if not coverage_str:
        return []

    coverages = []
    # 간단한 파싱: 쉼표로 분리하고 각각을 "이름 금액" 형식으로 처리
    items = coverage_str.split(",")
    for item in items:
        item = item.strip()
        if not item:
            continue

        # 숫자 부분 추출
        match = re.search(r'(\d+)', item.replace(",", ""))
        if match:
            amount = int(match.group(1))
            name = re.sub(r'\d+[만원]*', '', item).strip()
            if name:
                coverages.append({"name": name, "amount": amount})

    return coverages


class ReportParseError(Exception):
    """Excel 파싱 오류."""
    pass
